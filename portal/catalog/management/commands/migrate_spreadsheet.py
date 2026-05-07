"""
Management command para migrar a planilha v3.1 (504 estudos, 4 abas BDU) para o Nou-Rau.

Uso:
    python manage.py migrate_spreadsheet /caminho/para/planilha.xlsx
    python manage.py migrate_spreadsheet /caminho/para/planilha.xlsx --dry-run
    python manage.py migrate_spreadsheet /caminho/para/planilha.xlsx --sheet "Eventos"
"""

import os
import re
import unicodedata

import openpyxl
import psycopg2
from django.core.management.base import BaseCommand, CommandError


# Abas de dados na planilha v3.1 (ordem de processamento)
DATA_SHEETS = ["Eventos", "Livros Digitais", "Trabalhos Acadêmicos", "Materiais Pedagógicos"]

# Mapeamento de colunas da planilha v3.1 para campos internos.
# As chaves são os nomes EXATOS dos cabeçalhos da planilha.
COLUMN_MAP = {
    "Assunto": "assunto",
    "Categoria": "categoria",
    "Subcategoria": "subcategoria",
    "Microcategoria": "microcategoria",
    "Coleção": "colecao",
    "Tipo de informação": "tipo_informacao",
    "Autor Principal": "author",
    "Título Principal": "title",
    "Título variante/outro idioma": "title_en",
    "Autoridade Intelectual": "autor_principal",
    "Assunto em português": "keywords",
    "Assunto em outro idioma": "keywords_en",
    "Resumo": "abstract",
    "Abstract": "abstract_en",
    "Nota": "description",
    "Edição": "edicao",
    "Apresentações do Evento": "event_description",
    "Imprenta": "source",
    "Descrição Física": "descricao_fisica",
    "ISSN/ISBN": "nlspi",
    "Identificador do Objeto Digital (DOI)": "doi",
    "Acesso Eletrônico": "acesso_eletronico",
    "Inserir uma Capa": "capa",
    "Permissão de acesso ao material": "tacesso_raw",
    "Tipologia": "tipologia",
    "Complexidade": "complexidade",
    "Aplicabilidade": "uso_futuro",
    "Método": "metodo",
    "Resultados": "resultado",
    "Referências": "referencias",
    "Ano": "year",
}


def strip_accents(s):
    """Remove acentos de uma string para comparação fuzzy."""
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")


def normalize_text(val):
    """Limpa e normaliza texto da planilha."""
    if val is None:
        return ""
    return str(val).strip()


def extract_year(val):
    """Extrai ano de um valor."""
    val = normalize_text(val)
    match = re.search(r"\b(19|20)\d{2}\b", val)
    return match.group(0) if match else val


def generate_code(sequence_num):
    """Gera código único para o documento."""
    return f"bdlp-{sequence_num:06d}"


class Command(BaseCommand):
    help = "Migra planilha v3.1 de estudos (4 abas BDU) para o banco Nou-Rau"

    def add_arguments(self, parser):
        parser.add_argument("spreadsheet", type=str, help="Caminho para a planilha .xlsx")
        parser.add_argument("--dry-run", action="store_true", help="Apenas simula, sem inserir dados")
        parser.add_argument("--sheet", type=str, default=None, help="Nome de uma aba específica")
        parser.add_argument(
            "--skip-red", action="store_true",
            help="Pula linhas marcadas com fundo vermelho (curadoria sinaliza materiais a remover).",
        )
        parser.add_argument(
            "--start-seq", type=int, default=1,
            help="Sequência inicial dos códigos bdlp-XXXXXX (default 1). "
                 "Use para imports incrementais sem colidir com códigos existentes.",
        )

    def _get_write_connection(self):
        """Cria conexão de escrita usando o usuário php (não o portal_reader)."""
        return psycopg2.connect(
            dbname=os.environ.get("POSTGRES_DB", "nourau"),
            user=os.environ.get("POSTGRES_USER", "php"),
            password=os.environ.get("POSTGRES_PASSWORD", "abc123"),
            host=os.environ.get("POSTGRES_HOST", "postgres"),
            port=os.environ.get("POSTGRES_PORT", "5432"),
        )

    def handle(self, *args, **options):
        spreadsheet_path = options["spreadsheet"]
        dry_run = options["dry_run"]
        sheet_name = options.get("sheet")
        skip_red = options.get("skip_red", False)
        start_seq = options.get("start_seq", 1)

        # read_only=False para ter acesso a fill (cor de fundo) — só carrega
        # com formatting quando precisamos detectar células vermelhas.
        try:
            wb = openpyxl.load_workbook(spreadsheet_path, read_only=not skip_red)
        except Exception as e:
            raise CommandError(f"Erro ao abrir planilha: {e}")

        # Determinar quais abas processar
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise CommandError(f"Aba '{sheet_name}' não encontrada. Disponíveis: {wb.sheetnames}")
            sheets_to_process = [sheet_name]
        else:
            sheets_to_process = [s for s in DATA_SHEETS if s in wb.sheetnames]
            if not sheets_to_process:
                raise CommandError(f"Nenhuma aba de dados encontrada. Disponíveis: {wb.sheetnames}")

        self.stdout.write(f"Abas a processar: {sheets_to_process}")

        # Detecção de linhas vermelhas (curadoria marca para remover/excluir)
        red_rows_by_sheet = {}
        if skip_red:
            self.stdout.write("Detectando linhas vermelhas (curadoria)...")
            for sn in sheets_to_process:
                red_rows_by_sheet[sn] = self._detect_red_rows(wb[sn])
                if red_rows_by_sheet[sn]:
                    self.stdout.write(f"  [{sn}] {len(red_rows_by_sheet[sn])} linhas vermelhas serão puladas")

        # Carregar mapeamentos do banco
        conn = self._get_write_connection()
        try:
            topic_map = self._load_topic_map(conn)
            category_map = self._load_category_map(conn)
            type_info_map = self._load_type_information_map(conn)
            assunto_map = self._load_assunto_map(conn)
            subcategoria_map = self._load_subcategoria_map(conn)
            microcategoria_map = self._load_microcategoria_map(conn)

            self.stdout.write(f"Coleções raiz: {[k for k, v in topic_map.items() if v['parent_id'] == 0]}")
            self.stdout.write(f"Categorias: {len(category_map)}")
            self.stdout.write(f"Tipos de informação: {len(type_info_map)}")
            self.stdout.write(
                f"Taxonomia: {len(assunto_map)} assuntos, "
                f"{len(subcategoria_map)} subcategorias, "
                f"{len(microcategoria_map)} microcategorias"
            )

            global_seq = start_seq  # Contador global de código (continua import incremental)
            total_inserted = 0
            total_skipped = 0
            total_errors = []

            for sheet in sheets_to_process:
                ws = wb[sheet]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append(row)

                if len(rows) < 2:
                    self.stdout.write(self.style.WARNING(f"\n[{sheet}] Aba vazia, pulando."))
                    continue

                # Detectar layout do template "Inserir Material":
                #   L1 = seções ('CLASSIFICAÇÃO' / 'METADADOS BDU')
                #   L2 = nomes das colunas (Assunto, Categoria, ...)
                #   L3 = indicadores '✱ OBRIGATÓRIO' (não é dado)
                #   L4+ = dados
                # Para BDU: L1 = headers, L2+ = dados (defaults).
                if sheet == "Inserir Material":
                    header_idx = 1   # 0-based: row 2 da planilha
                    data_start = 3   # 0-based: row 4 da planilha
                else:
                    header_idx = 0
                    data_start = 1

                # Mapear cabeçalhos
                header = [normalize_text(h) for h in rows[header_idx]]
                col_indices = self._map_headers(header)

                mapped = len(col_indices)
                total_cols = len(COLUMN_MAP)
                self.stdout.write(f"\n[{sheet}] Colunas mapeadas: {mapped}/{total_cols}")
                if missing := set(COLUMN_MAP.values()) - set(col_indices.keys()):
                    # Filtrar opcionais que não existem em todas as abas
                    optional = {"edicao", "event_description", "nlspi"}
                    real_missing = missing - optional
                    if real_missing:
                        self.stdout.write(self.style.WARNING(f"  Colunas não encontradas: {real_missing}"))

                # Filtrar linhas com dados (mantém row_num original p/ skip-red)
                red_rows = red_rows_by_sheet.get(sheet, set())
                data_rows = []
                for offset, row in enumerate(rows[data_start:], start=data_start):
                    excel_row = offset + 1  # 1-based: linha real na planilha
                    if not any(v is not None for v in row):
                        continue
                    if excel_row in red_rows:
                        continue  # curadoria marcou pra remover
                    data_rows.append((excel_row, row))

                red_skipped = len(red_rows) if red_rows else 0
                self.stdout.write(
                    f"  Registros com dados: {len(data_rows)}"
                    + (f" (+ {red_skipped} vermelhos pulados)" if red_skipped else "")
                )

                sheet_inserted = 0
                sheet_skipped = 0
                sheet_errors = []

                for row_num, row in data_rows:
                    try:
                        record = self._parse_row(row, col_indices)
                        if not record.get("title"):
                            sheet_skipped += 1
                            continue

                        topic_id = self._resolve_topic(record, topic_map)
                        category_id = self._resolve_category(record.get("categoria", ""), category_map)
                        type_info_id = self._resolve_type_information(
                            record.get("tipo_informacao", ""), type_info_map
                        )
                        assunto_id = self._resolve_assunto(
                            record.get("assunto", ""), assunto_map
                        )
                        subcategoria_id = self._resolve_subcategoria(
                            record.get("subcategoria", ""), category_id, subcategoria_map
                        )
                        microcategoria_id = self._resolve_microcategoria(
                            record.get("microcategoria", ""), subcategoria_id, microcategoria_map
                        )

                        if dry_run:
                            title_preview = record["title"][:80]
                            self.stdout.write(f"  [DRY] #{global_seq} {sheet}/{row_num}: {title_preview}")
                            sheet_inserted += 1
                            global_seq += 1
                            continue

                        code = generate_code(global_seq)
                        self._insert_document(
                            conn, record, code,
                            topic_id, category_id, type_info_id,
                            assunto_id, subcategoria_id, microcategoria_id,
                        )
                        sheet_inserted += 1
                        global_seq += 1

                    except Exception as e:
                        sheet_errors.append((sheet, row_num, str(e)))
                        if len(sheet_errors) <= 5:
                            self.stdout.write(self.style.ERROR(f"  ERRO {sheet}/L{row_num}: {e}"))

                self.stdout.write(f"  [{sheet}] Inseridos: {sheet_inserted}, Ignorados: {sheet_skipped}, Erros: {len(sheet_errors)}")
                total_inserted += sheet_inserted
                total_skipped += sheet_skipped
                total_errors.extend(sheet_errors)

            if not dry_run:
                conn.commit()

        finally:
            conn.close()

        wb.close()

        self.stdout.write(self.style.SUCCESS(f"\n=== Resultado Final ==="))
        self.stdout.write(f"  Total inseridos: {total_inserted}")
        self.stdout.write(f"  Total ignorados: {total_skipped}")
        self.stdout.write(f"  Total erros: {len(total_errors)}")

        if total_errors:
            self.stdout.write(self.style.ERROR("\nPrimeiros 20 erros:"))
            for sheet, row_num, err in total_errors[:20]:
                self.stdout.write(self.style.ERROR(f"  {sheet}/L{row_num}: {err}"))

        if dry_run:
            self.stdout.write(self.style.WARNING("\n[DRY RUN] Nenhum dado foi inserido."))

    def _detect_red_rows(self, ws):
        """Retorna set com row_num (1-based) das linhas que contêm pelo menos
        uma célula com fundo vermelho (FF****00 ou similar).

        Usado para cumprir solicitação de curadoria: marcar materiais como
        vermelhos na planilha para que sejam excluídos do import.
        Verifica as primeiras 20 colunas (suficiente para Assunto, Título, etc.).
        """
        red_rows = set()
        max_check_col = min(ws.max_column, 20)
        for row_idx in range(2, ws.max_row + 1):  # pula header
            for col_idx in range(1, max_check_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                fill = cell.fill
                if fill is None or fill.fgColor is None:
                    continue
                rgb = fill.fgColor.rgb
                if not isinstance(rgb, str) or len(rgb) < 6:
                    continue
                rgb6 = rgb[-6:].upper()
                if not all(c in "0123456789ABCDEF" for c in rgb6):
                    continue
                r = int(rgb6[0:2], 16)
                g = int(rgb6[2:4], 16)
                b = int(rgb6[4:6], 16)
                # vermelho dominante: R alto, G/B baixos
                if r > 150 and g < 100 and b < 100:
                    red_rows.add(row_idx)
                    break
        return red_rows

    def _map_headers(self, header):
        """Mapeia cabeçalhos da planilha para campos internos com tolerância a acentos."""
        col_indices = {}
        header_normalized = [strip_accents(h.lower()) if h else "" for h in header]

        for col_name, field_name in COLUMN_MAP.items():
            col_norm = strip_accents(col_name.lower())
            for i, h_norm in enumerate(header_normalized):
                if h_norm and col_norm in h_norm:
                    col_indices[field_name] = i
                    break
        return col_indices

    def _parse_row(self, row, col_indices):
        """Extrai e normaliza campos de uma linha da planilha."""
        record = {}
        for field, idx in col_indices.items():
            record[field] = normalize_text(row[idx]) if idx < len(row) else ""
        return record

    def _load_topic_map(self, conn):
        """Carrega mapeamento nome→id de tópicos do banco."""
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, name, parent_id FROM topic")
            rows = cursor.fetchall()
        return {row[1].strip().lower(): {"id": row[0], "parent_id": row[2], "name": row[1]} for row in rows}

    def _load_category_map(self, conn):
        """Carrega mapeamento nome→id de categorias do banco."""
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, name FROM nr_category")
            rows = cursor.fetchall()
        return {row[1].strip().lower(): row[0] for row in rows}

    def _load_type_information_map(self, conn):
        """Carrega mapeamento nome→id de tipos de informação."""
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, name FROM type_information")
            rows = cursor.fetchall()
        result = {}
        for row in rows:
            if row[1]:
                result[strip_accents(row[1].strip().lower())] = row[0]
        return result

    def _load_assunto_map(self, conn):
        """Mapeia nome (lower, sem acento)→id em nr_assunto."""
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, nome FROM nr_assunto")
            rows = cursor.fetchall()
        return {strip_accents(r[1].strip().lower()): r[0] for r in rows if r[1]}

    def _load_subcategoria_map(self, conn):
        """Mapeia (category_id, nome_lower_sem_acento)→id em nr_subcategoria."""
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, nome, category_id FROM nr_subcategoria")
            rows = cursor.fetchall()
        return {
            (r[2], strip_accents(r[1].strip().lower())): r[0]
            for r in rows if r[1]
        }

    def _load_microcategoria_map(self, conn):
        """Mapeia (subcategoria_id, nome_lower_sem_acento)→id em nr_microcategoria."""
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, nome, subcategoria_id FROM nr_microcategoria")
            rows = cursor.fetchall()
        return {
            (r[2], strip_accents(r[1].strip().lower())): r[0]
            for r in rows if r[1]
        }

    def _resolve_topic(self, record, topic_map):
        """Resolve o topic_id em duas etapas:
        1) Coleção raiz pela coluna 'Coleção' (parent_id == 0).
        2) Subcoleção sob a raiz, casando 'Tipo de informação' com o nome da
           subcoleção (normalização de acento + tolerância a plural).

        Se não houver subcoleção compatível, retorna o ID da raiz.
        """
        # 1) Resolve raiz
        colecao_nome = record.get("colecao", "").strip().lower()
        root_id = None
        if colecao_nome and colecao_nome in topic_map:
            info = topic_map[colecao_nome]
            if info["parent_id"] == 0:
                root_id = info["id"]

        if root_id is None:
            # Fallback: primeira coleção raiz disponível
            for info in topic_map.values():
                if info["parent_id"] == 0:
                    root_id = info["id"]
                    break
            if root_id is None:
                return None

        # 2) Tenta refinar para subcoleção via Tipo de informação
        tipo_info = record.get("tipo_informacao", "").strip()
        if tipo_info:
            sub_id = self._match_subcollection(tipo_info, root_id, topic_map)
            if sub_id is not None:
                return sub_id

        return root_id

    def _match_subcollection(self, tipo_info_nome, root_id, topic_map):
        """Acha subcoleção sob root_id cujo nome casa com tipo_info_nome.

        Estratégias em ordem (primeira que casar vence):
          a) match exato normalizado (sem acento, lowercase)
          b) plural-tolerante (compara strip('s') dos dois lados)
          c) substring com >= 5 chars (para casos como "Capítulo de livro" → "Livro")
        """
        target = strip_accents(tipo_info_nome.lower()).strip()
        if not target:
            return None

        target_singular = target.rstrip("s")

        candidates = [
            (info["id"], strip_accents(info["name"].lower()).strip())
            for info in topic_map.values()
            if info["parent_id"] == root_id
        ]

        # a) Match exato
        for tid, name in candidates:
            if name == target:
                return tid

        # b) Plural-tolerante
        for tid, name in candidates:
            if name.rstrip("s") == target_singular:
                return tid

        # c) Substring (mín. 5 chars do candidato)
        for tid, name in candidates:
            if len(name) >= 5 and (name in target or target in name):
                return tid

        return None

    def _resolve_category(self, categoria_name, category_map):
        """Resolve o category_id baseado no nome da categoria.
        Retorna None quando não encontra — evita assignar a documentos uma
        categoria-fallback "lixo" (ex.: id 7 antigo) que polui as facetas.
        """
        if not categoria_name:
            return None
        key = strip_accents(categoria_name.strip().lower())
        for cat_key, cat_id in category_map.items():
            if strip_accents(cat_key) == key:
                return cat_id
        # Busca parcial — só se substring é >= 8 chars para evitar
        # falsos positivos como "PCA" matching "PCA EM ANEXO".
        if len(key) >= 8:
            for cat_key, cat_id in category_map.items():
                cat_norm = strip_accents(cat_key)
                if key in cat_norm or cat_norm in key:
                    return cat_id
        return None

    def _resolve_assunto(self, name, assunto_map):
        """Resolve assunto_id por nome (case e accent insensitive)."""
        if not name:
            return None
        key = strip_accents(name.strip().lower())
        return assunto_map.get(key)

    def _resolve_subcategoria(self, name, category_id, sub_map):
        """Resolve subcategoria_id sob category_id; faz match parcial se necessário."""
        if not name or not category_id:
            return None
        key = strip_accents(name.strip().lower())
        sub_id = sub_map.get((category_id, key))
        if sub_id:
            return sub_id
        # Match parcial dentro da mesma categoria
        for (cid, nome_norm), sid in sub_map.items():
            if cid != category_id:
                continue
            if key in nome_norm or nome_norm in key:
                return sid
        return None

    def _resolve_microcategoria(self, name, subcategoria_id, mic_map):
        """Resolve microcategoria_id sob subcategoria_id."""
        if not name or not subcategoria_id:
            return None
        key = strip_accents(name.strip().lower())
        mic_id = mic_map.get((subcategoria_id, key))
        if mic_id:
            return mic_id
        for (sid, nome_norm), mid in mic_map.items():
            if sid != subcategoria_id:
                continue
            if key in nome_norm or nome_norm in key:
                return mid
        return None

    def _resolve_type_information(self, tipo_info_name, type_info_map):
        """Resolve type_information id com fallback para NULL."""
        if not tipo_info_name:
            return None
        key = strip_accents(tipo_info_name.strip().lower())
        if key in type_info_map:
            return type_info_map[key]
        # Busca parcial para variações como "Artigo de periódico" -> "Artigos"
        # Mapeamento manual de variações conhecidas
        aliases = {
            "artigo de periodico": "artigos",
            "dissertacao": "dissertacoes",
            "tese": "teses",
            "tcc": "tccs",
            "monografia de especializacao": "dissertacoes",
            "livro": "livros",
            "capitulo de livro": "capitulo de livro",
            "artigo de evento": "artigos de eventos",
            "apresentacao de evento": "apresentacoes",
            "relatorio": "relatorios",
            "material pedagogico": "apostilas",
            "pagina web": None,  # não existe no Nou-Rau
            "documento normativo": None,
            "nota tecnica": "nota tecnica",
            "publicacao digital": None,
            "policy brief": "policy brief",
        }
        alias = aliases.get(key)
        if alias and alias in type_info_map:
            return type_info_map[alias]
        return None  # Sem match — será gravado como nome cru no campo info

    def _insert_document(
        self, conn, record, code,
        topic_id, category_id, type_info_id,
        assunto_id=None, subcategoria_id=None, microcategoria_id=None,
    ):
        """Insere um documento no banco via SQL direto."""
        # Montar description/info com campos auxiliares
        info_parts = []
        if record.get("nota"):
            info_parts.append(record["nota"])
        if record.get("tipo_informacao") and type_info_id is None:
            info_parts.append(f"Tipo de informação: {record['tipo_informacao']}")
        info = "\n".join(info_parts)

        # Autor principal: usar Autoridade Intelectual, fallback para Autor Principal
        autor_principal = record.get("autor_principal", "")
        if not autor_principal:
            autor_principal = record.get("author", "")

        # Etapa do processo licitatório: usar Subcategoria como rótulo livre (compatibilidade)
        etapa = record.get("subcategoria", "")[:255] if record.get("subcategoria") else ""

        # Ano: extrair como inteiro dedicado (preferir coluna `ano`)
        year_raw = record.get("year", "")
        year_str = extract_year(year_raw)
        try:
            ano_int = int(year_str) if year_str and year_str.isdigit() else None
        except (TypeError, ValueError):
            ano_int = None

        # Imprenta/source: se vazio, usar Ano
        source = record.get("source", "")
        if not source and year_str:
            source = year_str

        # Permissão: 'Aberto' (default), 'Restrito', ou texto cru se outro valor
        permissao_raw = (record.get("tacesso_raw") or "").strip()
        if permissao_raw.lower().startswith("aberto"):
            permissao = "Aberto"
        elif permissao_raw.lower().startswith("restrito"):
            permissao = "Restrito"
        else:
            permissao = permissao_raw[:20] if permissao_raw else None

        with conn.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO nr_document (
                    title, title_en, author, autor_principal,
                    abstract, abstract_en, keywords, keywords_en,
                    code, topic_id, category_id, status, remote,
                    acesso_eletronico, doi, nlspi, source,
                    tipologia, etapa_processo_licitatorio, complexidade,
                    uso_futuro, metodo, resultado, referencias,
                    description, info, owner_id,
                    typeinformation, typeinform_id,
                    descricao_fisica, edicao, event_description, capa,
                    assunto_id, subcategoria_id, microcategoria_id,
                    ano, permissao
                ) VALUES (
                    %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, 'a', 'y',
                    %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, 1,
                    %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s
                )
                """,
                [
                    record.get("title", "")[:1500],
                    record.get("title_en", "")[:1500],
                    record.get("author", "")[:3000],
                    autor_principal[:800],
                    record.get("abstract", ""),
                    record.get("abstract_en", ""),
                    record.get("keywords", ""),
                    record.get("keywords_en", ""),
                    code,
                    topic_id,
                    category_id,
                    record.get("acesso_eletronico", "")[:800],
                    record.get("doi", "")[:180],
                    record.get("nlspi", "")[:200],
                    source[:900],
                    record.get("tipologia", "")[:255],
                    etapa,
                    record.get("complexidade", "")[:50],
                    record.get("uso_futuro", ""),
                    record.get("metodo", ""),
                    record.get("resultado", ""),
                    record.get("referencias", ""),
                    "",  # description
                    info,
                    type_info_id,  # typeinformation (INT FK) — NULL if no match
                    type_info_id,  # typeinform_id (INT FK) — same
                    record.get("descricao_fisica", "")[:500],
                    record.get("edicao", "")[:900],
                    record.get("event_description", "")[:5000],
                    record.get("capa", "")[:65],
                    assunto_id,
                    subcategoria_id,
                    microcategoria_id,
                    ano_int,
                    permissao,
                ],
            )
