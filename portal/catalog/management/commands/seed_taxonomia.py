"""
Popula as tabelas nr_assunto, nr_subcategoria, nr_microcategoria a partir do
template oficial BDLP_Template_Insercao.xlsx (abas "REF" e "Árvore de Classificação").

Garante também:
  - Categorias (nr_category) cobrindo as 8 macroetapas do template
  - Complexidade "Média-Alta" (4º nível) presente em nr_complexidade

Uso:
    python manage.py seed_taxonomia /caminho/para/BDLP_Template_Insercao.xlsx
    python manage.py seed_taxonomia /caminho/para/BDLP_Template_Insercao.xlsx --dry-run
"""

import os
import re
import unicodedata

import openpyxl
import psycopg2
from django.core.management.base import BaseCommand, CommandError


def slugify(text):
    text = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode("ascii")
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def normalize(val):
    if val is None:
        return ""
    return str(val).strip()


class Command(BaseCommand):
    help = "Seed da taxonomia (Assunto/Subcategoria/Microcategoria) do template oficial"

    def add_arguments(self, parser):
        parser.add_argument("template", type=str, help="Caminho do BDLP_Template_Insercao.xlsx")
        parser.add_argument("--dry-run", action="store_true", help="Apenas reporta, não grava")

    def _connect(self):
        return psycopg2.connect(
            dbname=os.environ.get("POSTGRES_DB", "nourau"),
            user=os.environ.get("POSTGRES_USER", "php"),
            password=os.environ.get("POSTGRES_PASSWORD", "abc123"),
            host=os.environ.get("POSTGRES_HOST", "postgres"),
            port=os.environ.get("POSTGRES_PORT", "5432"),
        )

    def handle(self, *args, **options):
        path = options["template"]
        dry_run = options["dry_run"]

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f"Não foi possível abrir o template: {e}")

        if "REF" not in wb.sheetnames or "Árvore de Classificação" not in wb.sheetnames:
            raise CommandError(
                f"Template inválido. Abas esperadas: 'REF', 'Árvore de Classificação'. "
                f"Encontradas: {wb.sheetnames}"
            )

        assuntos = self._read_assuntos(wb["REF"])
        categorias = self._read_categorias(wb["REF"])
        arvore = self._read_arvore(wb["Árvore de Classificação"])
        wb.close()

        self.stdout.write(f"Assuntos no template: {len(assuntos)}")
        self.stdout.write(f"Categorias no template: {len(categorias)}")
        self.stdout.write(f"Linhas da árvore (categoria→subcategoria→microcategoria): {len(arvore)}")

        if dry_run:
            self.stdout.write(self.style.WARNING("[DRY RUN] Nenhuma alteração será gravada."))
            self.stdout.write("\n=== Assuntos ===")
            for a in assuntos:
                self.stdout.write(f"  - {a}")
            self.stdout.write("\n=== Categorias ===")
            for c in categorias:
                self.stdout.write(f"  - {c}")
            self.stdout.write("\n=== Árvore ===")
            for cat, sub, mic in arvore:
                self.stdout.write(f"  - {cat} > {sub or '·'} > {mic or '·'}")
            return

        conn = self._connect()
        try:
            # 1. Garantir Complexidade "Média-Alta" (template usa 4 níveis)
            self._ensure_complexidade(conn)

            # 2. Garantir Categorias em nr_category
            cat_ids = self._upsert_categorias(conn, categorias)

            # 3. Inserir Assuntos
            self._upsert_assuntos(conn, assuntos)

            # 4. Inserir Subcategorias e Microcategorias
            sub_inserted, mic_inserted = self._upsert_arvore(conn, arvore, cat_ids)

            conn.commit()
        finally:
            conn.close()

        self.stdout.write(self.style.SUCCESS("\n=== Seed concluído ==="))
        self.stdout.write(f"  Assuntos inseridos/atualizados: {len(assuntos)}")
        self.stdout.write(f"  Categorias garantidas: {len(categorias)}")
        self.stdout.write(f"  Subcategorias inseridas: {sub_inserted}")
        self.stdout.write(f"  Microcategorias inseridas: {mic_inserted}")

    # --- Leitura do template ---

    def _read_assuntos(self, ws):
        """Coluna 'ASSUNTOS' (índice 0) na aba REF, ignorando header."""
        items = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            v = normalize(row[0]) if row else ""
            if v:
                items.append(v)
        # Remover duplicatas preservando ordem
        seen = set()
        unique = []
        for v in items:
            k = v.lower()
            if k not in seen:
                seen.add(k)
                unique.append(v)
        return unique

    def _read_categorias(self, ws):
        """Coluna 'CATEGORIAS' (índice 1) na aba REF."""
        items = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            v = normalize(row[1]) if row and len(row) > 1 else ""
            if v:
                items.append(v)
        seen = set()
        unique = []
        for v in items:
            k = v.lower()
            if k not in seen:
                seen.add(k)
                unique.append(v)
        return unique

    def _read_arvore(self, ws):
        """Linhas (categoria, subcategoria, microcategoria) da aba 'Árvore de Classificação'."""
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            cat = normalize(row[0]) if row else ""
            sub = normalize(row[1]) if len(row) > 1 else ""
            mic = normalize(row[2]) if len(row) > 2 else ""
            if cat:
                rows.append((cat, sub or None, mic or None))
        return rows

    # --- Operações no banco ---

    def _ensure_complexidade(self, conn):
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO nr_complexidade (nome) VALUES ('Média-Alta') ON CONFLICT (nome) DO NOTHING"
            )

    def _upsert_categorias(self, conn, categorias):
        """Garante que cada categoria existe em nr_category. Retorna {nome_lower: id}."""
        result = {}
        with conn.cursor() as cur:
            cur.execute("SELECT id, name FROM nr_category")
            for cid, name in cur.fetchall():
                if name:
                    result[name.strip().lower()] = cid

            for cat in categorias:
                key = cat.lower()
                if key in result:
                    continue
                cur.execute(
                    "INSERT INTO nr_category (name, description, max_size) VALUES (%s, %s, 0) "
                    "ON CONFLICT (name) DO UPDATE SET description = EXCLUDED.description "
                    "RETURNING id",
                    [cat[:50], cat[:150]],
                )
                new_id = cur.fetchone()[0]
                result[key] = new_id
                self.stdout.write(f"  + Categoria: {cat}")
        return result

    def _upsert_assuntos(self, conn, assuntos):
        with conn.cursor() as cur:
            for ordem, nome in enumerate(assuntos, start=1):
                cur.execute(
                    "INSERT INTO nr_assunto (nome, slug, ordem) VALUES (%s, %s, %s) "
                    "ON CONFLICT (nome) DO UPDATE SET ordem = EXCLUDED.ordem",
                    [nome, slugify(nome), ordem],
                )

    def _upsert_arvore(self, conn, arvore, cat_ids):
        """Insere subcategorias e microcategorias respeitando a hierarquia."""
        sub_count = 0
        mic_count = 0
        sub_index = {}  # (category_id, subcat_lower) -> sub_id

        with conn.cursor() as cur:
            # Carregar subcategorias existentes para reuso
            cur.execute("SELECT id, nome, category_id FROM nr_subcategoria")
            for sid, nome, cid in cur.fetchall():
                sub_index[(cid, nome.strip().lower())] = sid

            # Pré-popular: subcategoria == nome igual à categoria conta como "raiz" da categoria
            for ordem, (cat, sub, mic) in enumerate(arvore, start=1):
                cat_id = cat_ids.get(cat.lower())
                if not cat_id:
                    self.stdout.write(self.style.WARNING(f"  ! Categoria não resolvida: {cat}"))
                    continue

                if sub:
                    sub_key = (cat_id, sub.lower())
                    if sub_key not in sub_index:
                        cur.execute(
                            "INSERT INTO nr_subcategoria (nome, slug, category_id, ordem) "
                            "VALUES (%s, %s, %s, %s) "
                            "ON CONFLICT (slug, category_id) DO UPDATE SET ordem = EXCLUDED.ordem "
                            "RETURNING id",
                            [sub, slugify(sub), cat_id, ordem],
                        )
                        sub_index[sub_key] = cur.fetchone()[0]
                        sub_count += 1

                if mic:
                    sub_key = (cat_id, sub.lower()) if sub else None
                    sub_id = sub_index.get(sub_key) if sub_key else None
                    if not sub_id:
                        # Microcategoria sem subcategoria explícita — pula com aviso
                        self.stdout.write(
                            self.style.WARNING(f"  ! Microcategoria sem sub: {cat}>{sub}>{mic}")
                        )
                        continue
                    cur.execute(
                        "INSERT INTO nr_microcategoria (nome, slug, subcategoria_id, ordem) "
                        "VALUES (%s, %s, %s, %s) "
                        "ON CONFLICT (slug, subcategoria_id) DO UPDATE SET ordem = EXCLUDED.ordem",
                        [mic, slugify(mic), sub_id, ordem],
                    )
                    mic_count += 1

        return sub_count, mic_count
